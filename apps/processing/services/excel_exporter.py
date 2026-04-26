"""Exportación Excel para corridas procesadas.

El exportador serializa los depósitos en un formato de entrega estable para la
UI y para descarga directa del usuario.
"""

from io import BytesIO

from django.core.files.base import ContentFile
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from apps.processing.models import ProcessRun

HEADERS = [
    "SECUENCIA",
    "FECHA CONSIGNACION",
    "HORA CONSIGNACION",
    "REFERENCIA",
    "VALOR",
    "ARCHIVO ORIGEN",
    "ES MES ACTUAL",
    "OBSERVACIONES",
]


def export_job_to_excel(process_run):
    """Genera y persiste el Excel asociado a una corrida procesada.

    Side Effects:
        Reemplaza el archivo exportado previo, escribe un nuevo blob en storage
        y actualiza el modelo `ProcessRun`.
    """
    process_run = ProcessRun.objects.prefetch_related("source_images__deposits").get(
        pk=process_run.pk
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Consignaciones"
    sheet.append(HEADERS)
    _style_headers(sheet[1])
    for source_image in process_run.source_images.order_by("sequence_index", "id"):
        for deposit in source_image.deposits.order_by("sequence_index", "id"):
            sheet.append(
                [
                    deposit.sequence_index,
                    deposit.fecha_consignacion,
                    deposit.hora_consignacion,
                    deposit.referencia,
                    float(deposit.valor),
                    source_image.source_name,
                    "SI" if deposit.is_current_month else "NO",
                    " | ".join(deposit.observations),
                ]
            )
    for column in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(max_length + 2, 40)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        row[4].number_format = '"$"#,##0.00'
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"job_{process_run.pk}_consignaciones.xlsx"
    if process_run.excel_file:
        process_run.excel_file.delete(save=False)
    process_run.excel_file.save(filename, ContentFile(output.read()), save=False)
    process_run.save(update_fields=["excel_file", "updated_at"])
    return process_run


def _style_headers(header_row):
    """Aplica formato visual al encabezado exportado."""
    fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    for cell in header_row:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
