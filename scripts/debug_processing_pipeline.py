#!/usr/bin/env python
from __future__ import annotations

import io
import json
import os
import sys
import time
import zipfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "MCP_back.settings")
os.environ.setdefault("STUB_PROVIDERS", "1")

import django

django.setup()

from django.conf import settings
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import load_workbook
from rest_framework.test import APIClient

from apps.processing.models import ExtractionLog, ProcessRun, SourceImage


def build_deposit_png(reference: str, value: str, date: str, hour: str) -> bytes:
    from PIL import Image, ImageDraw, ImageFont

    image = Image.new("RGB", (900, 360), "white")
    draw = ImageDraw.Draw(image)
    try:
        font_title = ImageFont.truetype("DejaVuSans-Bold.ttf", 36)
        font_body = ImageFont.truetype("DejaVuSans.ttf", 32)
    except OSError:
        font_title = ImageFont.load_default()
        font_body = ImageFont.load_default()
    draw.text((40, 35), "COMPROBANTE DE CONSIGNACION", fill="black", font=font_title)
    draw.text((40, 110), f"Referencia: {reference}", fill="black", font=font_body)
    draw.text((40, 165), f"Valor: {value}", fill="black", font=font_body)
    draw.text((40, 220), f"Fecha: {date}", fill="black", font=font_body)
    draw.text((40, 275), f"Hora: {hour}", fill="black", font=font_body)
    output = io.BytesIO()
    image.save(output, format="PNG")
    return output.getvalue()


def build_docx_with_images_and_text() -> bytes:
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w") as archive:
        archive.writestr(
            "[Content_Types].xml",
            """<?xml version="1.0" encoding="UTF-8"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Default Extension="png" ContentType="image/png"/>
  <Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>
</Types>""",
        )
        archive.writestr(
            "_rels/.rels",
            """<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>
</Relationships>""",
        )
        archive.writestr(
            "word/_rels/document.xml.rels",
            """<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rIdImage1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" Target="media/image1.png"/>
  <Relationship Id="rIdImage2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" Target="media/image2.png"/>
</Relationships>""",
        )
        archive.writestr(
            "word/document.xml",
            """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">
  <w:body>
    <w:p><w:r><w:t>Texto DOCX que debe guardarse como contexto, no como SourceImage.</w:t></w:r></w:p>
    <w:p><w:r><w:drawing><a:graphic><a:graphicData><pic:pic xmlns:pic="http://schemas.openxmlformats.org/drawingml/2006/picture"><pic:blipFill><a:blip r:embed="rIdImage1"/></pic:blipFill></pic:pic></a:graphicData></a:graphic></w:drawing></w:r></w:p>
    <w:p><w:r><w:drawing><a:graphic><a:graphicData><pic:pic xmlns:pic="http://schemas.openxmlformats.org/drawingml/2006/picture"><pic:blipFill><a:blip r:embed="rIdImage2"/></pic:blipFill></pic:pic></a:graphicData></a:graphic></w:drawing></w:r></w:p>
  </w:body>
</w:document>""",
        )
        archive.writestr(
            "word/media/image1.png",
            build_deposit_png("REF001", "50000", "15/04/2026", "09:30"),
        )
        archive.writestr(
            "word/media/image2.png",
            build_deposit_png("REF002", "75000", "16/04/2026", "10:45"),
        )
    buffer.seek(0)
    return buffer.getvalue()


def auth_headers() -> dict[str, str]:
    api_key = getattr(settings, "API_KEY", "") or "dev"
    return {"HTTP_X_API_KEY": api_key, "HTTP_HOST": "localhost"}


def assert_response(response, expected_status: int) -> dict:
    if response.status_code != expected_status:
        raise RuntimeError(
            f"Expected HTTP {expected_status}, got {response.status_code}: "
            f"{getattr(response, 'content', b'').decode(errors='replace')}"
        )
    return response.json()


def poll_until_terminal(client: APIClient, job_id: int, timeout_seconds: int = 30) -> dict:
    deadline = time.monotonic() + timeout_seconds
    latest = {}
    while time.monotonic() < deadline:
        latest = assert_response(client.get(f"/api/jobs/{job_id}/", **auth_headers()), 200)
        if latest["status"] in {"completed", "completed_with_errors", "failed"}:
            return latest
        time.sleep(0.5)
    raise TimeoutError(f"Job {job_id} did not reach a terminal state: {latest}")


def main() -> int:
    client = APIClient()
    upload = SimpleUploadedFile(
        "debug_pipeline.docx",
        build_docx_with_images_and_text(),
        content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    )
    uploaded = assert_response(
        client.post(
            "/api/documents/upload/",
            {"file": upload},
            format="multipart",
            **auth_headers(),
        ),
        201,
    )
    job_id = uploaded["id"]
    process_response = client.post(f"/api/jobs/{job_id}/process/", **auth_headers())
    if process_response.status_code not in {200, 202}:
        raise RuntimeError(process_response.content.decode(errors="replace"))
    final_detail = (
        process_response.json()
        if process_response.json()["status"] in {"completed", "completed_with_errors", "failed"}
        else poll_until_terminal(client, job_id)
    )
    logs = assert_response(client.get(f"/api/jobs/{job_id}/logs/", **auth_headers()), 200)
    exported = assert_response(
        client.post(f"/api/jobs/{job_id}/export/", **auth_headers()), 200
    )

    job = ProcessRun.objects.prefetch_related("source_images__deposits").get(pk=job_id)
    real_images = list(
        SourceImage.objects.filter(process_run=job).exclude(image_file="").order_by("sequence_index")
    )
    empty_image_sources = list(
        SourceImage.objects.filter(process_run=job, image_file="").values_list(
            "sequence_index", "source_name"
        )
    )
    excel_rows = 0
    if job.excel_file:
        with job.excel_file.open("rb") as workbook_file:
            workbook = load_workbook(workbook_file)
            excel_rows = max(workbook.active.max_row - 1, 0)

    summary = {
        "job_id": job_id,
        "status": final_detail["status"],
        "source_docx_exists": bool(job.source_docx),
        "extracted_text_len": len(job.extracted_text or ""),
        "total_images": job.total_images,
        "real_source_images": len(real_images),
        "empty_image_sources": empty_image_sources,
        "total_records": job.total_records,
        "db_deposits": job.deposits.count(),
        "api_deposits": sum(len(image["deposits"]) for image in final_detail["source_images"]),
        "ocr_logs": ExtractionLog.objects.filter(process_run=job, stage="ocr_extracted").count(),
        "llm_logs": ExtractionLog.objects.filter(process_run=job, stage="llm_structured").count(),
        "logs_count": len(logs),
        "excel_file": exported["excel_file"],
        "excel_rows": excel_rows,
        "source_images": [
            {
                "id": image.pk,
                "sequence_index": image.sequence_index,
                "source_name": image.source_name,
                "has_image_file": bool(image.image_file),
                "ocr_status": image.ocr_status,
                "deposits": image.deposits.count(),
            }
            for image in real_images
        ],
    }
    print(json.dumps(summary, indent=2, sort_keys=True))
    if summary["status"] != "completed":
        raise RuntimeError("Expected completed status")
    if summary["real_source_images"] != 2 or summary["empty_image_sources"]:
        raise RuntimeError("Expected exactly two real SourceImage rows and no empty image source")
    if summary["total_records"] <= 0 or summary["excel_rows"] != summary["total_records"]:
        raise RuntimeError("Expected persisted deposits and matching Excel rows")
    if summary["ocr_logs"] != 2 or summary["llm_logs"] != 2:
        raise RuntimeError("Expected one OCR and one LLM structuring call per real image")
    return 0


if __name__ == "__main__":
    sys.exit(main())
